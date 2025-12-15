from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from app.db.database import get_db
from app.db import models, schemas
from app.core.security import get_password_hash
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

router = APIRouter()

# --- NOVO ENDPOINT: LISTAR ESCALÕES ---
@router.get("/aux/escaloes")
def get_escaloes_list(db: Session = Depends(get_db)):
    """Retorna a lista de escalões para o dropdown do frontend."""
    return db.query(models.Escalao).all()

# --- LISTAR DEPARTAMENTOS ---
@router.get("/aux/departamentos")
def get_departamentos_list(db: Session = Depends(get_db)):
    """Retorna a lista de departamentos para o frontend."""
    return db.query(models.Departamento).all()

# --- FUNÇÕES AUXILIARES (Para encontrar IDs a partir de nomes) ---

def find_departamento_id(db: Session, nome_dept: str) -> Optional[int]:
    """Procura o ID do departamento. Tenta match exato ou parcial."""
    if not nome_dept or pd.isna(nome_dept): 
        return None
    
    nome_dept = str(nome_dept).strip()
    
    # 1. Tenta exato
    dept = db.query(models.Departamento).filter(models.Departamento.Nome == nome_dept).first()
    if dept: 
        return dept.Depart_id
    
    # 2. Tenta parcial
    dept = db.query(models.Departamento).filter(models.Departamento.Nome.ilike(f"%{nome_dept}%")).first()
    if dept: 
        return dept.Depart_id
    
    return None

def find_escalao_id(db: Session, nome_esc: str) -> Optional[int]:
    """Procura o ID do escalão. Flexível para 'Esc 1' ou '1.º Escalão'."""
    if not nome_esc or pd.isna(nome_esc): 
        return None
    
    nome_esc = str(nome_esc).strip()
    
    # 1. Tenta match exato
    esc = db.query(models.Escalao).filter(models.Escalao.Nome == nome_esc).first()
    if esc: 
        return esc.Escalao_id
    
    # 2. Tenta extrair o número (ex: "1.º Escalão" -> 1, "Esc 1" -> 1)
    numero = ''.join(filter(str.isdigit, str(nome_esc)))
    if numero:
        # Procura na base de dados algum escalão que contenha esse número
        esc = db.query(models.Escalao).filter(models.Escalao.Nome.like(f"%{numero}%")).first()
        if esc: 
            return esc.Escalao_id
        
    return None

# --- CREATE ---
@router.post("/", response_model=schemas.StaffListagem)
def create_staff(staff: schemas.StaffCreate, db: Session = Depends(get_db)):
    # 1. Validação de Salário
    if staff.Salario and staff.Salario > 999999.99:
        raise HTTPException(status_code=400, detail="Vencimento demasiado grande")

    # 2. Verificar se email já existe
    email_exists_staff = db.query(models.Staff).filter(models.Staff.email == staff.email).first()
    email_exists_prof = db.query(models.Professor).filter(models.Professor.email == staff.email).first()
    
    if email_exists_staff or email_exists_prof:
        raise HTTPException(status_code=400, detail="Email já registado")

    # 3. Criar conforme a Role
    if staff.role == "teacher":
        # Resolver IDs das relações usando as funções auxiliares
        dept_id = find_departamento_id(db, staff.Departamento)
        esc_id = find_escalao_id(db, staff.Escalao)

        new_prof = models.Professor(
            Nome=staff.Nome,
            email=staff.email,
            hashed_password=get_password_hash("123mudar"),
            Telefone=staff.Telefone,
            Morada=staff.Morada,
            Data_Nasc="1980-01-01",
            Depart_id=dept_id, # Associa ID
            Escalao_id=esc_id,  # Associa ID
            role="teacher"
        )
        try:
            db.add(new_prof)
            db.commit()
            db.refresh(new_prof)
            
            # Buscar valor do salário através da relação para devolver ao frontend
            salario_real = 0.0
            if new_prof.escalao:
                salario_real = float(new_prof.escalao.Valor_Base)

            return {
                "Staff_id": new_prof.Professor_id,
                "Nome": new_prof.Nome,
                "email": new_prof.email,
                "Cargo": "Docente",
                "role": "teacher",
                "Telefone": new_prof.Telefone,
                "Morada": new_prof.Morada,
                "Salario": salario_real,
                "Escalao": new_prof.escalao.Nome if new_prof.escalao else None,
                "Departamento": new_prof.departamento.Nome if new_prof.departamento else None
            }
        except Exception as e:
            db.rollback()
            print(f"Erro BD: {e}")
            raise HTTPException(status_code=500, detail=f"Erro ao criar professor: {str(e)}")

    else:
        # Staff normal
        new_staff = models.Staff(
            Nome=staff.Nome,
            email=staff.email,
            hashed_password=get_password_hash("123mudar"),
            Telefone=staff.Telefone,
            Morada=staff.Morada,
            Cargo=staff.Cargo,
            role=staff.role if staff.role else "staff",
            Salario=staff.Salario,
            Escalao=staff.Escalao
        )
        try:
            db.add(new_staff)
            db.commit()
            db.refresh(new_staff)
            return new_staff
        except Exception as e:
            db.rollback()
            print(f"Erro BD: {e}")
            raise HTTPException(status_code=500, detail=f"Erro ao criar staff: {str(e)}")

# --- UPDATE ---
@router.put("/{staff_id}", response_model=schemas.StaffListagem)
def update_staff(staff_id: int, staff_data: schemas.StaffUpdate, db: Session = Depends(get_db)):
    
    if staff_data.Salario is not None and staff_data.Salario > 999999.99:
        raise HTTPException(status_code=400, detail="Vencimento demasiado grande")

    if staff_data.role == "teacher":
        prof = db.query(models.Professor).filter(models.Professor.Professor_id == staff_id).first()
        if not prof:
            raise HTTPException(status_code=404, detail="Professor não encontrado")
        
        if staff_data.Nome: prof.Nome = staff_data.Nome
        if staff_data.email: prof.email = staff_data.email
        if staff_data.Telefone: prof.Telefone = staff_data.Telefone
        if staff_data.Morada: prof.Morada = staff_data.Morada
        
        # Atualizar Relações (Lookup de IDs)
        if staff_data.Departamento:
            prof.Depart_id = find_departamento_id(db, staff_data.Departamento)
        
        if staff_data.Escalao:
            prof.Escalao_id = find_escalao_id(db, staff_data.Escalao)
        
        try:
            db.commit()
            db.refresh(prof)
            
            salario_real = 0.0
            if prof.escalao: salario_real = float(prof.escalao.Valor_Base)

            return {
                "Staff_id": prof.Professor_id,
                "Nome": prof.Nome,
                "email": prof.email,
                "Cargo": "Docente",
                "role": "teacher",
                "Telefone": prof.Telefone,
                "Morada": prof.Morada,
                "Salario": salario_real,
                "Escalao": prof.escalao.Nome if prof.escalao else None,
                "Departamento": prof.departamento.Nome if prof.departamento else None
            }
        except Exception as e:
            db.rollback()
            raise HTTPException(status_code=500, detail=f"Erro ao atualizar: {str(e)}")

    else:
        staff = db.query(models.Staff).filter(models.Staff.Staff_id == staff_id).first()
        if not staff:
            raise HTTPException(status_code=404, detail="Staff não encontrado")

        if staff_data.Nome: staff.Nome = staff_data.Nome
        if staff_data.email: staff.email = staff_data.email
        if staff_data.Telefone: staff.Telefone = staff_data.Telefone
        if staff_data.Morada: staff.Morada = staff_data.Morada
        if staff_data.Cargo: staff.Cargo = staff_data.Cargo
        
        if staff_data.Salario is not None: staff.Salario = staff_data.Salario
        if staff_data.Escalao is not None: staff.Escalao = staff_data.Escalao

        try:
            db.commit()
            db.refresh(staff)
            return staff
        except Exception as e:
            db.rollback()
            raise HTTPException(status_code=500, detail=f"Erro ao atualizar: {str(e)}")

# --- READ ---
@router.get("/", response_model=List[schemas.StaffDisplay])
def read_staff(skip: int = 0, limit: int = 1000, db: Session = Depends(get_db)):
    try:
        staff_query = db.query(models.Staff).all()
        # Carregar as relações para eficiência
        prof_query = db.query(models.Professor)\
            .options(joinedload(models.Professor.escalao), joinedload(models.Professor.departamento))\
            .all()
            
        results = []

        for s in staff_query:
            results.append({
                "id": s.Staff_id,
                "email": s.email,
                "Nome": s.Nome,
                "Cargo": s.Cargo,
                "role": getattr(s, "role", "staff"),
                "Telefone": s.Telefone,
                "Morada": s.Morada,
                "Salario": s.Salario,
                "Escalao": s.Escalao,
                "Departamento": None
            })

        for p in prof_query:
            # Ler valores das tabelas relacionadas
            salario_valor = float(p.escalao.Valor_Base) if p.escalao else 0.0
            escalao_nome = p.escalao.Nome if p.escalao else None
            dept_nome = p.departamento.Nome if p.departamento else None

            results.append({
                "id": p.Professor_id,
                "email": p.email,
                "Nome": p.Nome,
                "Cargo": f"Docente {dept_nome}" if dept_nome else "Docente",
                "role": "teacher",
                "Telefone": p.Telefone,
                "Morada": p.Morada,
                "Salario": salario_valor,
                "Escalao": escalao_nome,
                "Departamento": dept_nome
            })

        results.sort(key=lambda x: x["Nome"])
        return results[skip : skip + limit]

    except Exception as e:
        print("Erro ao ler lista:", e)
        raise HTTPException(status_code=500, detail="Erro ao ler lista")

# --- DELETE ---
@router.delete("/{id}", status_code=204)
def delete_staff_member(id: int, role: Optional[str] = "staff", db: Session = Depends(get_db)):
    member = None
    if role == "teacher":
        member = db.query(models.Professor).filter(models.Professor.Professor_id == id).first()
    else:
        member = db.query(models.Staff).filter(models.Staff.Staff_id == id).first()

    if not member:
        raise HTTPException(status_code=404, detail="Membro não encontrado")
    
    try:
        if role == "teacher":
            # Limpar dependências para evitar erro de FK
            turmas_diretor = db.query(models.Turma).filter(models.Turma.DiretorT == id).all()
            for t in turmas_diretor: t.DiretorT = None
            db.query(models.TurmaDisciplina).filter(models.TurmaDisciplina.Professor_id == id).delete()
            db.query(models.Ocorrencia).filter(models.Ocorrencia.Professor_id == id).delete()
            db.delete(member)
        else:
            db.delete(member)

        db.commit()
    
    except Exception as e:
        db.rollback()
        print(f"Erro ao eliminar: {e}")
        raise HTTPException(status_code=400, detail=f"Não foi possível eliminar: {str(e)}")
    
    return

# --- IMPORT / EXPORT / TEMPLATE (INTELIGENTE) ---

@router.get("/data/template")
def get_staff_template(db: Session = Depends(get_db)):
    """
    Gera um Excel com duas folhas:
    1. 'Importar Dados': Com dropdowns e formatação.
    2. 'Guia e Regras': Com instruções e listas de referência.
    """
    
    # 1. Buscar Listas da BD
    depts = [d.Nome for d in db.query(models.Departamento).all()]
    escaloes = [e.Nome for e in db.query(models.Escalao).all()]
    
    wb = Workbook()
    
    # --- FOLHA 1: IMPORTAR DADOS ---
    ws_main = wb.active
    ws_main.title = "Importar Dados"
    
    headers = ["Nome", "Email", "Telefone", "Morada", "Role", "Cargo", "Salario", "Escalao", "Departamento"]
    ws_main.append(headers)
    
    # Estilo Cabeçalho
    for cell in ws_main[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # Dados de Exemplo
    example_data = [
        ["Ex: Maria Santos", "maria@escola.pt", "910000000", "Rua A", "staff", "Secretária", 950.00, "", ""],
        ["Ex: Prof João", "joao@escola.pt", "920000000", "Rua B", "teacher", "", 0, escaloes[0] if escaloes else "Esc 1", depts[0] if depts else "Ciências"]
    ]
    for row in example_data:
        ws_main.append(row)

    # --- FOLHA 2: GUIA E REGRAS ---
    ws_rules = wb.create_sheet("Guia e Regras")
    
    # Tabela de Regras
    ws_rules.append(["Campo", "Obrigatório?", "Regras / Instruções"])
    regras = [
        ["Nome", "Sim", "Nome completo do funcionário."],
        ["Email", "Sim", "Deve ser único no sistema e conter '@'."],
        ["Telefone", "Não", "9 dígitos (preferencial)."],
        ["Morada", "Não", "Texto livre."],
        ["Role", "Sim", "Selecione: 'staff' (Funcionário), 'teacher' (Professor) ou 'admin'."],
        ["Cargo", "Sim (Staff)", "Ex: Secretário, Porteiro. Ignorado para Professores."],
        ["Salario", "Sim (Staff)", "Valor numérico (Ex: 1200.50). Máx: 999999.99. Ignorado para Professores."],
        ["Escalao", "Sim (Prof)", "Selecione o Escalão da lista."],
        ["Departamento", "Sim (Prof)", "Selecione o Departamento da lista."]
    ]
    for r in regras:
        ws_rules.append(r)
        
    # Estilo Tabela Regras
    for cell in ws_rules[1]:
        cell.font = Font(bold=True)
    
    # --- LISTAS DE REFERÊNCIA (Para os Dropdowns) ---
    # Escrevemos as listas na folha de regras, colunas E e F
    ws_rules["E1"] = "Departamentos Válidos"
    ws_rules["F1"] = "Escalões Válidos"
    ws_rules["E1"].font = Font(bold=True)
    ws_rules["F1"].font = Font(bold=True)

    for i, d in enumerate(depts, start=2):
        ws_rules[f"E{i}"] = d
    
    for i, e in enumerate(escaloes, start=2):
        ws_rules[f"F{i}"] = e

    # --- DATA VALIDATION (DROPDOWNS) ---

    
    # 1. Role Dropdown (Coluna E)
    dv_role = DataValidation(type="list", formula1='"staff,teacher,admin"', allow_blank=True)
    dv_role.error = "Por favor selecione staff, teacher ou admin."
    dv_role.errorTitle = "Entrada Inválida"
    ws_main.add_data_validation(dv_role)
    dv_role.add("E2:E200") # Aplica às primeiras 200 linhas

    # 2. Escalão Dropdown (Coluna H) - Referência à folha de regras
    if escaloes:
        last_row = len(escaloes) + 1
        formula_esc = f"'Guia e Regras'!$F$2:$F${last_row}"
        dv_esc = DataValidation(type="list", formula1=formula_esc, allow_blank=True)
        ws_main.add_data_validation(dv_esc)
        dv_esc.add("H2:H200")

    # 3. Departamento Dropdown (Coluna I)
    if depts:
        last_row = len(depts) + 1
        formula_dept = f"'Guia e Regras'!$E$2:$E${last_row}"
        dv_dept = DataValidation(type="list", formula1=formula_dept, allow_blank=True)
        ws_main.add_data_validation(dv_dept)
        dv_dept.add("I2:I200")

    # Ajustar larguras das colunas
    for col in range(1, 10):
        ws_main.column_dimensions[get_column_letter(col)].width = 20
    ws_rules.column_dimensions['A'].width = 15
    ws_rules.column_dimensions['C'].width = 50
    ws_rules.column_dimensions['E'].width = 25
    ws_rules.column_dimensions['F'].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers={"Content-Disposition": "attachment; filename=template_staff_inteligente.xlsx"}
    )

@router.get("/data/export")
def export_staff_data(db: Session = Depends(get_db)):
    """Exporta para Excel usando a biblioteca openpyxl para manter a consistência."""
    # Reutiliza a lógica de leitura
    lista = read_staff(skip=0, limit=10000, db=db)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados Exportados"
    
    # Headers
    headers = ["Nome", "Email", "Telefone", "Morada", "Role", "Cargo", "Salario", "Escalao", "Departamento"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    # Rows
    for item in lista:
        # Item pode ser dict (do read_staff) ou objeto. read_staff devolve dicts.
        row = [
            item.get("Nome"), 
            item.get("email"), 
            item.get("Telefone"), 
            item.get("Morada"),
            item.get("role"), 
            item.get("Cargo"), 
            item.get("Salario"),
            item.get("Escalao"), 
            item.get("Departamento")
        ]
        ws.append(row)
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=staff_export.xlsx"})

@router.post("/data/import")
async def import_staff_data(file: UploadFile = File(...), db: Session = Depends(get_db)):
    try:
        contents = await file.read()
        # openpyxl engine para ler ficheiros excel modernos
        df = pd.read_excel(io.BytesIO(contents), engine='openpyxl')
        df = df.where(pd.notnull(df), None)
        
        count_success = 0
        for index, row in df.iterrows():
            try:
                # 1. Ignorar Ex e vazios
                nome = str(row.get("Nome", ""))
                if not nome or nome.startswith("Ex:") or nome == "None": 
                    continue
                
                email = str(row.get("Email", ""))
                if not email or "@" not in email: 
                    continue
                
                # 2. Check Duplicados
                if db.query(models.Staff).filter_by(email=email).first() or \
                   db.query(models.Professor).filter_by(email=email).first(): 
                    continue

                # 3. Role Map
                role_input = str(row.get("Role", "staff")).lower()
                
                if "prof" in role_input or "teacher" in role_input: 
                    role = "teacher"
                elif "admin" in role_input: 
                    role = "admin"
                else: 
                    role = "staff"
                
                # 4. Dados
                tel = str(row.get("Telefone", "")) if row.get("Telefone") else None
                morada = str(row.get("Morada", "")) if row.get("Morada") else None
                
                if role == "teacher":
                    new_prof = models.Professor(
                        Nome=nome, 
                        email=email, 
                        hashed_password=get_password_hash("123mudar"),
                        Telefone=tel, 
                        Morada=morada, 
                        Data_Nasc="1980-01-01", 
                        role="teacher",
                        Depart_id=find_departamento_id(db, row.get("Departamento")),
                        Escalao_id=find_escalao_id(db, row.get("Escalao"))
                    )
                    db.add(new_prof)
                else:
                    sal = 0.0
                    try: 
                        sal = float(row.get("Salario", 0))
                    except: pass
                    
                    if sal > 999999.99: 
                        continue # Skip if invalid salary
                    
                    new_staff = models.Staff(
                        Nome=nome, 
                        email=email, 
                        hashed_password=get_password_hash("123mudar"),
                        Telefone=tel, 
                        Morada=morada, 
                        Cargo=str(row.get("Cargo", "")),
                        role=role, 
                        Salario=sal, 
                        Escalao=str(row.get("Escalao", "")) if row.get("Escalao") else None
                    )
                    db.add(new_staff)
                
                count_success += 1
            except Exception as e:
                print(f"Erro linha {index}: {e}")
                continue
        
        db.commit()
        return {"message": f"Importados {count_success} registos."}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro fatal: {str(e)}")